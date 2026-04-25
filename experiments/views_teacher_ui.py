from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.hashers import make_password
from django.http import HttpResponse, JsonResponse
from django.utils.timezone import now
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Q
import csv
import io
import os
import re
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from django.conf import settings
from .experiment_tables import EXPERIMENT_TABLES
from experiments.models import StudentApproval
from accounts.models import User
from .models import (
    Batch,
    BatchExperiment,
    ShortAnswerBank,
    Exam,
    ExamMCQ,
    ExamShortAnswer,
    MCQBank,
    Experiment, 
    ExperimentAttempt,
    ExamSpotting,
    ExamPractical,
    SpottingBank,
    SessionalContinuousMark,
    YEAR_CHOICES
)

D_PHARM_NUMBERS = [1, 2, 3, 6, 8, 9, 10, 11, 16, 17, 20, 21, 24, 25, 26, 27, 28, 29, 31, 34]


# =====================================================
# TEACHER DASHBOARD (CARD VIEW ONLY)
# =====================================================
def teacher_dashboard(request):
    if request.session.get("role") != "teacher":
        return redirect("/login/")

    teacher = request.user

    students = User.objects.filter(
        role="student",
        created_by=teacher,
        college=teacher.college
    )

    subject_list = teacher.subject.split(",") if teacher.subject else []
    
    experiments = Experiment.objects.filter(is_active=True).order_by("number")

    # 🔥 Plan-based filtering
    college = teacher.college
    if college and college.selected_plan == 'dpharm':
        experiments = experiments.filter(number__in=D_PHARM_NUMBERS)

    attempts = ExperimentAttempt.objects.filter(
        student__created_by=teacher,
        experiment__is_active=True,
        completed_at__isnull=False
    )

    return render(
        request,
        "teacher/dashboard.html",
        {
            "name": teacher.first_name,
            "total_students": students.count(),
            "total_experiments": experiments.values('number').distinct().count(),
            "total_attempts": attempts.count(),
        }
    )


# =====================================================
# TEACHER → STUDENTS LIST
# =====================================================
def teacher_students(request):
    if request.session.get("role") != "teacher":
        return redirect("/login/")

    subject_code = request.GET.get("subject")
    
    students = User.objects.filter(
        role="student",
        created_by=request.user,
        college=request.user.college
    ).order_by("roll_no")

    if subject_code:
        students = students.filter(subject=subject_code)

    # Subject name for header
    subject_map = {
        "dpharm_2": "2ND YR D.PHARM",
        "bpharm_4": "2ND YR B.PHARM (SEM-IV)",
        "bpharm_5": "3RD YR B.PHARM (SEM-V)",
        "bpharm_6": "3RD YR B.PHARM (SEM-VI)"
    }
    subject_name = subject_map.get(subject_code, "All Students")

    # Fetch batches to map students
    from .models import Batch
    batches = Batch.objects.filter(teacher=request.user)
    
    # Map batches to students
    for student in students:
        student.assigned_batch_name = "---"
        if student.roll_no and student.roll_no.isdigit():
            roll_int = int(student.roll_no)
            for batch in batches:
                if batch.start_roll <= roll_int <= batch.end_roll:
                    student.assigned_batch_name = batch.name
                    break

    return render(
        request,
        "teacher/students.html",
        {
            "students": students,
            "subject_name": subject_name,
            "subject_code": subject_code
        }
    )


# =====================================================
# EXPORT STUDENTS → PDF (Branded)
# =====================================================
def export_students_pdf(request):
    if request.session.get("role") != "teacher":
        return redirect("/login/")

    subject_code = request.GET.get("subject")
    teacher = request.user
    college = teacher.college

    students = User.objects.filter(
        role="student",
        created_by=teacher,
        college=college
    ).order_by("roll_no")

    if subject_code:
        students = students.filter(subject=subject_code)

    # Batch mapping logic (same as teacher_students view)
    batches = Batch.objects.filter(teacher=teacher)
    for student in students:
        student.assigned_batch_name = "---"
        if student.roll_no and student.roll_no.isdigit():
            roll_int = int(student.roll_no)
            for b in batches:
                if b.start_roll <= roll_int <= b.end_roll:
                    student.assigned_batch_name = b.name
                    break

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=50)
    elements = []
    styles = getSampleStyleSheet()

    # --- Header (Logo and Name) ---
    header_data = []
    college_name = college.name if college else "Virtual Lab System"
    college_address = college.address if college and college.address else ""
    
    logo_part = None
    if college and college.logo:
        try:
            logo_path = college.logo.path
            if os.path.exists(logo_path):
                logo_part = Image(logo_path, width=65, height=65)
        except Exception:
            pass

    name_style = ParagraphStyle(
        'CollegeNameStyle',
        parent=styles['Normal'],
        fontName='Times-Bold',
        fontSize=22,
        textColor=colors.black,
        alignment=0, # Left
        leading=24
    )

    address_style = ParagraphStyle(
        'AddressStyle',
        parent=styles['Normal'],
        fontName='Times-Roman',
        fontSize=9,
        textColor=colors.grey,
        alignment=0, # Left
    )
    
    # Header Content (Name + Address)
    header_info = [Paragraph(college_name, name_style)]
    if college_address:
        header_info.append(Paragraph(college_address, address_style))

    if logo_part:
        # Table for Logo and info alignment
        header_table = Table([[logo_part, header_info]], colWidths=[80, 450])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ]))
        elements.append(header_table)
    else:
        elements.extend(header_info)
    
    elements.append(Spacer(1, 15))
    
    # Horizontal Line
    elements.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor("#ff4b2b"), spaceAfter=15))
    
    # Title "Student List" (Centered and Underlined)
    title_style = ParagraphStyle(
        'MainTitleStyle',
        parent=styles['Normal'],
        fontName='Times-Bold',
        fontSize=16,
        alignment=1, # Center
        spaceAfter=15,
    )
    elements.append(Paragraph("<u>STUDENT LIST</u>", title_style))

    subtitle_style = ParagraphStyle(
        'SubtitleStyle',
        parent=styles['Normal'],
        fontName='Times-Roman',
        fontSize=11,
        alignment=1, # Centered with title
        spaceAfter=20,
        textColor=colors.darkgrey
    )
    subject_map = {
        "dpharm_2": "2ND YR D.PHARM",
        "bpharm_4": "2ND YR B.PHARM (SEM-IV)",
        "bpharm_5": "3RD YR B.PHARM (SEM-V)",
        "bpharm_6": "3RD YR B.PHARM (SEM-VI)"
    }
    subtitle = f"Subject: {subject_map.get(subject_code, 'All Subjects')}"
    elements.append(Paragraph(subtitle, subtitle_style))

    # --- Table Data ---
    data = [["Roll No", "Batch", "Student Name", "Email ID", "Mobile", "Status"]]
    for s in students:
        status = "Active" if s.is_active else "Inactive"
        data.append([
            s.roll_no or "---",
            s.assigned_batch_name,
            f"{s.first_name} {s.last_name}",
            s.email,
            s.mobile or "---",
            status
        ])

    table = Table(data, colWidths=[55, 75, 135, 135, 75, 50]) # Total width = 525 (fits in 535 available)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#ff4b2b")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Times-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, 1), (-1, -1), 'Times-Roman'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#fff5f5")]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(table)

    # --- Page Setup (Border, Footer) ---
    def page_setup(canvas, doc):
        canvas.saveState()
        
        # --- Page Border ---
        canvas.setStrokeColor(colors.HexColor("#ff4b2b"))
        canvas.setLineWidth(1)
        canvas.rect(20, 20, A4[0]-40, A4[1]-40) # Standard page border
        
        # --- GMARS Logo and Footer ---
        try:
            gmars_logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'logo.png')
            if os.path.exists(gmars_logo_path):
                # Increased Height further
                canvas.drawImage(gmars_logo_path, A4[0] - 115, 1, width=85, height=90, mask='auto')
        except Exception:
            pass
        
        canvas.setFont('Times-BoldItalic', 10)
        canvas.drawRightString(A4[0] - 120, 40, "Signature of Examiner")
        
        canvas.setFont('Times-Roman', 8)
        canvas.drawString(40, 40, f"Generated on: {now().strftime('%d-%m-%Y %H:%M')}")
        canvas.drawCentredString(A4[0]/2, 40, f"Page {canvas.getPageNumber()}")
        canvas.restoreState()

    try:
        doc.build(elements, onFirstPage=page_setup, onLaterPages=page_setup)
    except Exception as e:
        return HttpResponse(f"Error generating PDF: {str(e)}", status=500)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Students_{subject_code or "All"}.pdf"'
    response.write(buffer.getvalue())
    buffer.close()
    return response


# =====================================================
# =====================================================
# PDF HELPERS
# =====================================================
def parse_html_to_reportlab_table(html_str, styles):
    """
    Parses a simple HTML table string from EXPERIMENT_TABLES 
    and returns a ReportLab Table object.
    It extracts headers and pre-fills the first column if data exists in <tbody>.
    """
    if not html_str:
        return None

    # Extract headers (between <th> and </th>)
    headers = re.findall(r'<th[^>]*>(.*?)</th>', html_str, re.IGNORECASE | re.DOTALL)
    if not headers:
        return None

    clean_headers = [re.sub(r'<[^>]+>', '', h).strip() for h in headers]
    data = [clean_headers]
    
    # Extract rows from <tbody>
    tbody_match = re.search(r'<tbody[^>]*>(.*?)</tbody>', html_str, re.IGNORECASE | re.DOTALL)
    if tbody_match:
        tbody_content = tbody_match.group(1)
        rows = re.findall(r'<tr[^>]*>(.*?)</tr>', tbody_content, re.IGNORECASE | re.DOTALL)
        
        for row_html in rows:
            cells = re.findall(r'<td[^>]*>(.*?)</td>', row_html, re.IGNORECASE | re.DOTALL)
            clean_cells = []
            for i, cell in enumerate(cells):
                # Prefill first column, keep others empty
                if i == 0:
                    clean_cells.append(re.sub(r'<[^>]+>', '', cell).strip())
                else:
                    clean_cells.append("")
            
            # Pad with empty strings if cells < headers
            while len(clean_cells) < len(clean_headers):
                clean_cells.append("")
            
            data.append(clean_cells)
    
    # If no rows found, add 5 empty ones
    if len(data) == 1:
        for _ in range(5):
            data.append(["" for _ in clean_headers])

    # Dynamic column widths
    cols = len(clean_headers)
    col_width = 535 / cols if cols > 0 else 100
    
    table = Table(data, colWidths=[col_width] * cols)
    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Times-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#6366f1")), # Premium Indigo Header
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, 1), (-1, -1), 'Times-Roman'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ]))
    
    return table

# =====================================================
# EXPORT EXAM QUESTIONS → PDF (Branded)
# =====================================================
def export_exam_questions_pdf(request, exam_id):
    if request.session.get("role") != "teacher":
        return redirect("/login/")

    exam = get_object_or_404(Exam, id=exam_id, teacher=request.user)
    college = request.user.college

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=50)
    elements = []
    styles = getSampleStyleSheet()

    # Branding Header (Replicated from export_students_pdf)
    college_name = college.name if college else "Virtual Lab System"
    college_address = college.address if college and college.address else ""
    
    logo_part = None
    if college and college.logo:
        try:
            logo_path = college.logo.path
            if os.path.exists(logo_path):
                logo_part = Image(logo_path, width=65, height=65)
        except Exception:
            pass

    name_style = ParagraphStyle('CollegeNameStyle', parent=styles['Normal'], fontName='Times-Bold', fontSize=22, textColor=colors.black, alignment=0, leading=24)
    address_style = ParagraphStyle('AddressStyle', parent=styles['Normal'], fontName='Times-Roman', fontSize=9, textColor=colors.grey, alignment=0)
    
    header_info = [Paragraph(college_name, name_style)]
    if college_address:
        header_info.append(Paragraph(college_address, address_style))

    # Teacher Attribution
    teacher_name = f"{request.user.first_name} {request.user.last_name}".strip() or request.user.username
    creator_style = ParagraphStyle('CreatorStyle', parent=styles['Normal'], fontName='Times-Bold', fontSize=10, textColor=colors.black, alignment=0)
    header_info.append(Paragraph(f"Created by: {teacher_name}", creator_style))

    if logo_part:
        header_table = Table([[logo_part, header_info]], colWidths=[80, 450])
        header_table.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), ('LEFTPADDING', (0, 0), (-1, -1), 0)]))
        elements.append(header_table)
    else:
        elements.extend(header_info)
    
    elements.append(Spacer(1, 15))
    elements.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor("#ff4b2b"), spaceAfter=15))

    # Exam Metadata Header
    meta_style = ParagraphStyle('MetaStyle', parent=styles['Normal'], fontName='Times-Bold', fontSize=12, alignment=1)
    elements.append(Paragraph(f"{exam.title} - {exam.get_year_display()}", meta_style))
    elements.append(Paragraph(f"Duration: {exam.duration_minutes} Minutes", meta_style))
    elements.append(Spacer(1, 20))

    # Question Paper Body
    q_title_style = ParagraphStyle('QTitle', parent=styles['Normal'], fontName='Times-Bold', fontSize=14, spaceAfter=10)
    text_style = ParagraphStyle('Text', parent=styles['Normal'], fontName='Times-Roman', fontSize=11, leading=14, spaceAfter=10)
    
    is_dpharm = exam.year == "dpharm_2"
    q_num = 1

    if is_dpharm:
        # =====================================================
        # D.PHARM PATTERN
        # =====================================================
        
        # Q1: Synopsis (MCQ + Short)
        mcqs = exam.mcqs.all()
        sas = exam.short_answers.all()
        if mcqs.exists() or sas.exists():
            total_synopsis_marks = sum(m.marks for m in mcqs) + sum(s.marks for s in sas)
            elements.append(Paragraph(f"Question {q_num}: Synopsis (A. MCQ, B. Short Answers) .................... ({total_synopsis_marks} Marks)", q_title_style))
            
            if mcqs.exists():
                elements.append(Paragraph("<u>A. Multiple Choice Questions</u>", ParagraphStyle('Sub', parent=styles['Normal'], fontName='Times-Bold', fontSize=12, spaceBefore=5, spaceAfter=8)))
                for i, mcq in enumerate(mcqs, 1):
                    elements.append(Paragraph(f"{i}. {mcq.question_text}", text_style))
                    elements.append(Paragraph(f"A) {mcq.option_a}", ParagraphStyle('Opt', parent=text_style, leftIndent=20, spaceAfter=2)))
                    elements.append(Paragraph(f"B) {mcq.option_b}", ParagraphStyle('Opt', parent=text_style, leftIndent=20, spaceAfter=2)))
                    elements.append(Paragraph(f"C) {mcq.option_c}", ParagraphStyle('Opt', parent=text_style, leftIndent=20, spaceAfter=2)))
                    elements.append(Paragraph(f"D) {mcq.option_d}", ParagraphStyle('Opt', parent=text_style, leftIndent=20, spaceAfter=10)))
            
            if sas.exists():
                elements.append(Paragraph("<u>B. Short Answer Questions</u>", ParagraphStyle('Sub', parent=styles['Normal'], fontName='Times-Bold', fontSize=12, spaceBefore=8, spaceAfter=8)))
                for i, sa in enumerate(sas, 1):
                    elements.append(Paragraph(f"{i}. {sa.question_text}", text_style))
            
            elements.append(Spacer(1, 15))
            q_num += 1

        # Q2: Experiment (Spotting, Major, Minor)
        spotting = exam.spotting_questions.all()
        major = exam.practicals.filter(practical_type="major").first()
        minor = exam.practicals.filter(practical_type="minor").first()
        
        if spotting.exists() or major or minor:
            elements.append(Paragraph(f"Question {q_num}: Experiment", q_title_style))
            
            # A. Spotting
            if spotting.exists():
                total_spot_marks = sum(s.marks for s in spotting)
                elements.append(Paragraph(f"<u>A. Spotting</u> ({total_spot_marks} Marks)", ParagraphStyle('Sub', parent=styles['Normal'], fontName='Times-Bold', fontSize=12, spaceBefore=5, spaceAfter=8)))
                
                spot_list_text = []
                for i, spot in enumerate(spotting, 1):
                    # Use the actual filename/slug instead of the descriptive name
                    if spot.bank_item and spot.bank_item.image_slug:
                        # Strip extension if it exists for a cleaner look (acto instead of acto.jpg)
                        slug_name = os.path.splitext(spot.bank_item.image_slug)[0]
                        name = slug_name
                    else:
                        name = "Untitled"
                    spot_list_text.append(f"<b>{i}.</b> {name} ({spot.marks} Marks)")
                
                # Join them with separate paragraphs
                for txt in spot_list_text:
                    elements.append(Paragraph(txt, text_style))
                    elements.append(Spacer(1, 5))
            
            # B. Major Experiment
            if major:
                elements.append(Paragraph(f"<u>B. Major Experiment</u> ({major.marks} Marks)", ParagraphStyle('Sub', parent=styles['Normal'], fontName='Times-Bold', fontSize=12, spaceBefore=5, spaceAfter=8)))
                elements.append(Paragraph(f"<b>Aim:</b> {major.aim}", text_style))
                elements.append(Spacer(1, 5))
            
            # C. Minor Experiment
            if minor:
                elements.append(Paragraph(f"<u>C. Minor Experiment</u> ({minor.marks} Marks)", ParagraphStyle('Sub', parent=styles['Normal'], fontName='Times-Bold', fontSize=12, spaceBefore=5, spaceAfter=8)))
                elements.append(Paragraph(f"<b>Aim:</b> {minor.aim}", text_style))
                elements.append(Spacer(1, 5))
            
            elements.append(Spacer(1, 15))
            q_num += 1

        # Q3: Viva Voce
        if exam.viva_marks > 0:
            elements.append(Paragraph(f"Question {q_num}: Viva Voce .................................................... ({exam.viva_marks} Marks)", q_title_style))
            elements.append(Spacer(1, 15))
            q_num += 1

        # Q4: Practical Record Maintenance
        if exam.practical_record_marks > 0:
            elements.append(Paragraph(f"Question {q_num}: Practical record maintenance ................................. ({exam.practical_record_marks} Marks)", q_title_style))
            elements.append(Spacer(1, 15))
            q_num += 1

    else:
        # =====================================================
        # STANDARD PATTERN (B.PHARM etc.)
        # =====================================================
        
        # Spotting (Usually not for B.Pharm but keeping logic if exists)
        spotting = exam.spotting_questions.all()
        if spotting.exists():
            total_marks = sum(s.marks for s in spotting)
            elements.append(Paragraph(f"Question {q_num}: Spotting ................................................. ({total_marks} Marks)", q_title_style))
            for i, spot in enumerate(spotting, 1):
                elements.append(Paragraph(f"<b>{i}</b> ({spot.marks} Marks)", text_style))
                if spot.bank_item and spot.bank_item.image_slug:
                    try:
                        img_path = os.path.join(settings.BASE_DIR, 'static', 'spotting_images', spot.bank_item.image_slug)
                        if os.path.exists(img_path): elements.append(Image(img_path, width=200, height=150))
                    except: pass
                elements.append(Spacer(1, 10))
            q_num += 1

        # Synopsis
        mcqs = exam.mcqs.all()
        sas = exam.short_answers.all()
        if mcqs.exists() or sas.exists():
            total_marks = sum(m.marks for m in mcqs) + sum(s.marks for s in sas)
            elements.append(Paragraph(f"Question {q_num}: Synopsis ................................................. ({total_marks} Marks)", q_title_style))
            if mcqs.exists():
                elements.append(Paragraph("<u>Multiple Choice Questions</u>", ParagraphStyle('Sub', parent=styles['Normal'], fontName='Times-Bold', fontSize=12, spaceAfter=8)))
                for i, mcq in enumerate(mcqs, 1):
                    elements.append(Paragraph(f"{i}. {mcq.question_text}", text_style))
                    elements.append(Paragraph(f"A) {mcq.option_a}, B) {mcq.option_b}, C) {mcq.option_c}, D) {mcq.option_d}", ParagraphStyle('Opt', parent=text_style, leftIndent=20)))
            if sas.exists():
                elements.append(Paragraph("<u>Short Answer Questions</u>", ParagraphStyle('Sub', parent=styles['Normal'], fontName='Times-Bold', fontSize=12, spaceBefore=8, spaceAfter=8)))
                for i, sa in enumerate(sas, 1):
                    elements.append(Paragraph(f"{i}. {sa.question_text}", text_style))
            q_num += 1

        # Practical
        major = exam.practicals.filter(practical_type="major").first()
        minor = exam.practicals.filter(practical_type="minor").first()
        if major or minor:
            elements.append(Paragraph(f"Question {q_num}: Experiments", q_title_style))
            if major:
                elements.append(Paragraph(f"<b>(A) Major Experiment</b> ({major.marks} Marks)", text_style))
                elements.append(Paragraph(f"Aim: {major.aim}", text_style))
            if minor:
                elements.append(Paragraph(f"<b>(B) Minor Experiment</b> ({minor.marks} Marks)", text_style))
                elements.append(Paragraph(f"Aim: {minor.aim}", text_style))
            q_num += 1

        # Viva
        if exam.viva_marks > 0:
            elements.append(Paragraph(f"Question {q_num}: Viva Voce .................................................... ({exam.viva_marks} Marks)", q_title_style))
        elements.append(Spacer(1, 10))

    # Page Footer Logic
    def page_setup(canvas, doc):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor("#ff4b2b"))
        canvas.setLineWidth(1)
        canvas.rect(20, 20, A4[0]-40, A4[1]-40)
        
        try:
            gmars_logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'logo.png')
            if os.path.exists(gmars_logo_path):
                canvas.drawImage(gmars_logo_path, A4[0] - 110, 15, width=80, height=80, mask='auto')
        except Exception: pass
        
        canvas.setFont('Times-BoldItalic', 10)
        canvas.drawRightString(A4[0] - 120, 40, "Signature of Examiner")
        canvas.setFont('Times-Roman', 8)
        canvas.drawString(40, 40, f"Generated on: {now().strftime('%d-%m-%Y %H:%M')}")
        canvas.drawCentredString(A4[0]/2, 40, f"Page {canvas.getPageNumber()}")
        canvas.restoreState()

    doc.build(elements, onFirstPage=page_setup, onLaterPages=page_setup)
    
    response = HttpResponse(content_type='application/pdf')
    filename = f"Question_Paper_{exam.title.replace(' ', '_')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write(buffer.getvalue())
    buffer.close()
    return response


# =====================================================
# EXPORT STUDENTS → EXCEL (Formatted)
# =====================================================
def export_students_excel(request):
    if request.session.get("role") != "teacher":
        return redirect("/login/")

    subject_code = request.GET.get("subject")
    teacher = request.user

    students = User.objects.filter(
        role="student",
        created_by=teacher,
        college=teacher.college
    ).order_by("roll_no")

    if subject_code:
        students = students.filter(subject=subject_code)

    # Batch mapping logic
    batches = Batch.objects.filter(teacher=teacher)
    for student in students:
        student.assigned_batch_name = "---"
        if student.roll_no and student.roll_no.isdigit():
            roll_int = int(student.roll_no)
            for b in batches:
                if b.start_roll <= roll_int <= b.end_roll:
                    student.assigned_batch_name = b.name
                    break

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students List"

    # Header Row
    headers = ["Roll No", "Batch", "Student Name", "Email ID", "Mobile No", "Status"]
    ws.append(headers)

    # Styling Header
    header_fill = PatternFill(start_color="FF4B2B", end_color="FF4B2B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment

    # Data Rows
    for s in students:
        ws.append([
            s.roll_no or "---",
            s.assigned_batch_name,
            f"{s.first_name} {s.last_name}",
            s.email,
            s.mobile or "---",
            "Active" if s.is_active else "Inactive"
        ])

    # Column Widths
    column_widths = [12, 15, 30, 35, 15, 12]
    for i, width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i+1)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Students_{subject_code or "All"}.xlsx"'
    return response


# =====================================================
# TEACHER → EXPERIMENTS LIST
# =====================================================
def teacher_experiments(request):
    if request.session.get("role") not in ["teacher", "superadmin"]:
        return redirect("/login/")

    experiments = Experiment.objects.filter(is_active=True).order_by("number")

    # 🔥 Plan-based filtering
    college = request.user.college
    if college and college.selected_plan == 'dpharm':
        experiments = experiments.filter(number__in=D_PHARM_NUMBERS)


    query = request.GET.get("q", "")
    if query:
        experiments = experiments.filter(
            Q(name__icontains=query) | Q(number__icontains=query)
        )

    return render(
        request,
        "teacher/experiments.html",
        {
            "experiments": experiments,
            "query": query
        }
    )


# =====================================================
# TEACHER → ATTEMPTS LIST
# =====================================================
def teacher_attempts(request):
    if request.session.get("role") != "teacher":
        return redirect("/login/")

    attempts = ExperimentAttempt.objects.filter(
        student__created_by=request.user,
        experiment__is_active=True,
        completed_at__isnull=False
    ).select_related(
        "student",
        "experiment"
    ).order_by("-completed_at")

    return render(
        request,
        "teacher/attempts.html",
        {
            "attempts": attempts
        }
    )


# =====================================================
# CREATE STUDENT (TEACHER → STUDENT)
# =====================================================
def create_student(request):
    if request.session.get("role") != "teacher":
        return redirect("/login/")

    teacher = request.user

    if request.method == "POST":
        name = request.POST.get("name")
        email = request.POST.get("email")
        password = request.POST.get("password")

        if not all([name, email, password]):
            return render(
                request,
                "teacher/create_student.html",
                {"error": "All fields are required"}
            )

        if User.objects.filter(email=email).exists():
            return render(
                request,
                "teacher/create_student.html",
                {"error": "Student with this email already exists"}
            )

        User.objects.create(
            username=email,
            email=email,
            first_name=name,
            role="student",
            college=teacher.college,
            created_by=teacher,
            password=make_password(password)
        )

        return redirect("/teacher/dashboard/")

    return render(request, "teacher/create_student.html")


# =====================================================
# CSV EXPORT → STUDENTS
# =====================================================
def teacher_export_students_csv(request):
    if request.session.get("role") != "teacher":
        return redirect("/login/")

    students = User.objects.filter(
        role="student",
        created_by=request.user,
        college=request.user.college
    )

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="students.csv"'

    writer = csv.writer(response)
    writer.writerow(["Name", "Email", "Active"])

    for s in students:
        writer.writerow([
            s.first_name,
            s.email,
            "Yes" if s.is_active else "No"
        ])

    return response


# =====================================================
# CSV EXPORT → ATTEMPTS
# =====================================================
def teacher_export_attempts_csv(request):
    if request.session.get("role") != "teacher":
        return redirect("/login/")

    attempts = ExperimentAttempt.objects.filter(
        student__created_by=request.user,
        experiment__is_active=True,
        completed_at__isnull=False
    )

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="attempts.csv"'

    writer = csv.writer(response)
    writer.writerow(["Student", "Experiment", "Completed At"])

    for a in attempts:
        writer.writerow([
            a.student.first_name,
            a.experiment.name,
            a.completed_at
        ])

    return response

# =====================================================
# 🔥 TEACHER → STUDENT APPROVAL REQUESTS (NEW FEATURE)
# =====================================================
def teacher_student_requests(request):
    if request.session.get("role") != "teacher":
        return redirect("/login/")

    pending_approvals = StudentApproval.objects.filter(
        selected_teacher=request.user,
        approval_status="pending"
    ).select_related("student")

    return render(
        request,
        "teacher/student_requests.html",
        {
            "pending_approvals": pending_approvals
        }
    )


# =====================================================
# 🔥 TEACHER → APPROVE STUDENT REQUEST
# =====================================================
def approve_student_request(request, student_id):
    if request.session.get("role") != "teacher":
        return redirect("/login/")

    approval = get_object_or_404(
        StudentApproval,
        student__id=student_id,
        selected_teacher=request.user,
        approval_status="pending"
    )

    # Update approval record
    approval.approval_status = "approved"
    approval.approved_by_teacher = True
    approval.save()

    # Activate student and assign subject
    student = approval.student
    student.is_active = True
    student.subject = approval.requested_subject
    student.save()

    return redirect("/teacher/student-requests/")


# =====================================================
# 🔥 TEACHER → APPROVE ALL STUDENT REQUESTS
# =====================================================
def approve_all_student_requests(request):
    if request.session.get("role") != "teacher":
        return redirect("/login/")

    pending_approvals = StudentApproval.objects.filter(
        selected_teacher=request.user,
        approval_status="pending"
    )

    for approval in pending_approvals:
        # Update approval record
        approval.approval_status = "approved"
        approval.approved_by_teacher = True
        approval.save()

        # Activate student and assign subject
        student = approval.student
        student.is_active = True
        student.subject = approval.requested_subject
        student.save()

    return redirect("/teacher/student-requests/")


# =====================================================
# 🔥 TEACHER → REJECT STUDENT REQUEST
# =====================================================
def reject_student_request(request, student_id):
    if request.session.get("role") != "teacher":
        return redirect("/login/")

    approval = get_object_or_404(
        StudentApproval,
        student__id=student_id,
        selected_teacher=request.user,
        approval_status="pending"
    )

    approval.approval_status = "rejected"
    approval.save()

    # Keep student inactive
    student = approval.student
    student.is_active = False
    student.save()

    return redirect("/teacher/student-requests/")

# =====================================================
# 🔥 TEACHER → MANAGE BATCHES
# =====================================================
def teacher_manage_batches(request):
    if request.session.get("role") != "teacher":
        return redirect("/login/")

    batches = Batch.objects.filter(
        teacher=request.user
    ).order_by("-created_at")

    return render(
        request,
        "teacher/manage_batches.html",
        {
            "batches": batches,
            "year_choices": YEAR_CHOICES
        }
    )


# =====================================================
# 🔥 TEACHER → CREATE BATCH
# =====================================================
def teacher_create_batch(request):
    if request.session.get("role") != "teacher":
        return redirect("/login/")

    if request.method == "POST":
        name = request.POST.get("name")
        start_roll = request.POST.get("start_roll")
        end_roll = request.POST.get("end_roll")

        if not all([name, start_roll, end_roll]):
            return redirect("/teacher/manage-batches/")

        batch = Batch.objects.create(
            teacher=request.user,
            name=name,
            start_roll=int(start_roll),
            end_roll=int(end_roll),
        )

        year = request.POST.get("year")
        from .models import BatchExtra
        BatchExtra.objects.create(batch=batch, year=year)

        return redirect("/teacher/manage-batches/")

    return redirect("/teacher/manage-batches/")


# =====================================================
# 🔥 TEACHER → EDIT BATCH
# =====================================================
def teacher_edit_batch(request, batch_id):
    if request.session.get("role") != "teacher":
        return redirect("/login/")

    batch = get_object_or_404(
        Batch,
        id=batch_id,
        teacher=request.user
    )

    if request.method == "POST":
        batch.name = request.POST.get("name")
        batch.start_roll = int(request.POST.get("start_roll"))
        batch.end_roll = int(request.POST.get("end_roll"))
        batch.save()

        year = request.POST.get("year")
        from .models import BatchExtra
        extra, created = BatchExtra.objects.get_or_create(batch=batch)
        extra.year = year
        extra.save()

        return redirect("/teacher/manage-batches/")

    return render(
        request,
        "teacher/edit_batch.html",
        {
            "batch": batch
        }
    )

# =====================================================
# 🔥 TEACHER → ASSIGN PRACTICAL TO BATCH
# =====================================================
def teacher_assign_practical(request, batch_id):
    if request.session.get("role") != "teacher":
        return redirect("/login/")

    batch = get_object_or_404(
        Batch,
        id=batch_id,
        teacher=request.user
    )

    experiments = Experiment.objects.filter(is_active=True).order_by("number")
    is_dpharm = False
    try:
        if hasattr(batch, 'extra') and batch.extra.year == 'dpharm_2':
            is_dpharm = True
    except:
        pass

    # 📂 Group experiments into Sections for cleaner rendering
    sections = [
        {"id": "A", "title": "Introduction & Basic Concepts", "exps": []},
        {"id": "B", "title": "Ethics, Handling & Techniques", "exps": []},
        {"id": "C", "title": "Autonomic & Peripheral Pharmacology", "exps": []},
        {"id": "D", "title": "Central Nervous System (CNS)", "exps": []},
        {"id": "E", "title": "Cardiovascular & Renal Pharmacology", "exps": []},
        {"id": "F", "title": "Analgesic & Anti-Inflammatory", "exps": []},
        {"id": "G", "title": "Endocrine & Metabolic Pharmacology", "exps": []},
        {"id": "H", "title": "Gastrointestinal & Anti-Ulcer", "exps": []},
        {"id": "I", "title": "Anaesthetics & Allergy Models", "exps": []},
        {"id": "J", "title": "In-Vitro Bioassay Studies", "exps": []},
        {"id": "K", "title": "Toxicology Studies", "exps": []},
        {"id": "L", "title": "Biomechanics & Biostatistics", "exps": []},
    ]

    for exp in experiments:
        num = exp.number
        if 1 <= num <= 6: sections[0]["exps"].append(exp)
        elif 7 <= num <= 9: sections[1]["exps"].append(exp)
        elif 10 <= num <= 15: sections[2]["exps"].append(exp)
        elif 16 <= num <= 22: sections[3]["exps"].append(exp)
        elif 23 <= num <= 29: sections[4]["exps"].append(exp)
        elif 30 <= num <= 33: sections[5]["exps"].append(exp)
        elif num == 34: sections[6]["exps"].append(exp)
        elif num == 35: sections[7]["exps"].append(exp)
        elif 36 <= num <= 40: sections[8]["exps"].append(exp)
        elif 41 <= num <= 50: sections[9]["exps"].append(exp)
        elif 51 <= num <= 53: sections[10]["exps"].append(exp)
        elif 54 <= num <= 56: sections[11]["exps"].append(exp)

    assigned_ids = list(
        BatchExperiment.objects.filter(batch=batch).values_list("experiment_id", flat=True)
    )

    if request.method == "POST":
        selected_ids = request.POST.getlist("experiments")
        BatchExperiment.objects.filter(batch=batch).delete()
        for eid in selected_ids:
            BatchExperiment.objects.create(batch=batch, experiment_id=eid)
        return redirect(f"/teacher/batch/{batch_id}/assign/")

    return render(
        request,
        "teacher/assign_practical.html",
        {
            "batch": batch,
            "sections": sections,
            "experiments": experiments,
            "assigned_ids": assigned_ids,
            "is_dpharm": is_dpharm
        }
    )

def teacher_exams(request):
    if request.session.get("role") != "teacher":
        return redirect("/login/")

    subject_code = request.GET.get("subject")
    
    exams = Exam.objects.filter(
        teacher=request.user
    )
    
    if subject_code:
        # Derive year from subject_code
        target_year = subject_code
        if subject_code in ["bpharm_5", "bpharm_6"]:
            target_year = "bpharm_56"
        exams = exams.filter(year=target_year)

    return render(
        request,
        "teacher/exams.html",
        {
            "exams": exams,
            "subject_code": subject_code
        }
    )

def teacher_create_exam(request):
    if request.session.get("role") != "teacher":
        return redirect("/login/")

    subject_code = request.GET.get("subject")

    if request.method == "POST":
        title = request.POST.get("title")
        exam_type = request.POST.get("exam_type")
        # year is now determined by subject context
        duration = request.POST.get("duration")

        # Derive year from subject_code
        # Mapping: dpharm_2 -> dpharm_2, bpharm_4 -> bpharm_4, bpharm_5/6 -> bpharm_56
        year = subject_code
        if subject_code in ["bpharm_5", "bpharm_6"]:
            year = "bpharm_56"

        if not all([title, exam_type, year, duration]):
            return redirect(f"/teacher/exams/?subject={subject_code or ''}")

        Exam.objects.create(
            teacher=request.user,
            title=title,
            exam_type=exam_type,
            year=year,
            duration_minutes=int(duration)
        )

        return redirect(f"/teacher/exams/?subject={subject_code or ''}")

    return render(request, "teacher/create_exam.html", {"subject_code": subject_code})

def teacher_edit_exam(request, exam_id):
    if request.session.get("role") != "teacher":
        return redirect("/login/")

    subject_code = request.GET.get("subject")

    exam = get_object_or_404(
        Exam,
        id=exam_id,
        teacher=request.user
    )

    if request.method == "POST":
        exam.title = request.POST.get("title")
        exam.exam_type = request.POST.get("exam_type")
        
        # Derived from subject context if provided, otherwise keep existing
        if subject_code:
            year = subject_code
            if subject_code in ["bpharm_5", "bpharm_6"]:
                year = "bpharm_56"
            exam.year = year
            
        exam.duration_minutes = int(request.POST.get("duration"))
        exam.save()

        return redirect(f"/teacher/exams/?subject={subject_code or ''}")

    return render(
        request,
        "teacher/edit_exam.html",
        {
            "exam": exam,
            "subject_code": subject_code
        }
    )

def teacher_toggle_exam(request, exam_id):
    if request.session.get("role") != "teacher":
        return redirect("/login/")

    subject_code = request.GET.get("subject")

    exam = get_object_or_404(
        Exam,
        id=exam_id,
        teacher=request.user
    )

    exam.is_active = not exam.is_active
    
    # 🔥 RESET LOGIC: If reactivating, clear all previous attempts
    # This ensures students start fresh (Start button instead of Resume)
    if exam.is_active:
        exam.attempts.all().delete()
        
    exam.save()

    return redirect(f"/teacher/exams/?subject={subject_code or ''}")

def teacher_exam_builder(request, exam_id):
    if request.session.get("role") != "teacher":
        return redirect("/login/")

    exam = get_object_or_404(
        Exam,
        id=exam_id,
        teacher=request.user
    )

    # ================================
    # HANDLE POST ACTIONS
    # ================================
    if request.method == "POST":

        action = request.POST.get("action")

        # --------------------------------
# ADD SPOTTING IMAGE (ONLY D.PHARM)
# --------------------------------
        if action == "add_spotting":

            # Restrict spotting to D.Pharm only
            if exam.year != "dpharm_2":
                return redirect(request.path)

            bank_id = request.POST.get("bank_id")
            bank_item = get_object_or_404(SpottingBank, id=bank_id)

            ExamSpotting.objects.create(
                exam=exam,
                bank_item=bank_item,
                marks=int(request.POST.get("marks", 1)),
                order=exam.spotting_questions.count() + 1
            )

            return redirect(request.path)
        
        # --------------------------------
        # UPDATE SPOTTING MARKS
        # --------------------------------
        elif action == "update_spotting_marks":

            if exam.year != "dpharm_2":
                return redirect(request.path)

            spot_id = request.POST.get("spot_id")
            marks = request.POST.get("marks")

            spot = get_object_or_404(
                ExamSpotting,
                id=spot_id, 
                exam=exam
            )

            spot.marks = int(marks)
            spot.save()

            return redirect(request.path)

        # --------------------------------
        # ADD MANUAL MCQ
        # --------------------------------
        elif action == "add_manual_mcq":

            ExamMCQ.objects.create(
                exam=exam,
                source_type="manual",
                question_text=request.POST.get("question_text"),
                option_a=request.POST.get("option_a"),
                option_b=request.POST.get("option_b"),
                option_c=request.POST.get("option_c"),
                option_d=request.POST.get("option_d"),
                correct_option=request.POST.get("correct_option"),
                marks=int(request.POST.get("marks", 1)),
                order=exam.mcqs.count() + 1
            )

            return redirect(request.path)

        # --------------------------------
        # ADD MANUAL SHORT ANSWER
        # --------------------------------
        elif action == "add_manual_short":

            ExamShortAnswer.objects.create(
                exam=exam,
                source_type="manual",
                question_text=request.POST.get("question_text"),
                marks=int(request.POST.get("marks", 5)),
                order=exam.short_answers.count() + 1
            )

            return redirect(request.path)

        # --------------------------------
        # ADD MCQ FROM BANK
        # --------------------------------
        elif action == "add_bank_mcq":

            selected_ids = request.POST.getlist("selected_mcqs")

            for q_id in selected_ids:
                bank_q = get_object_or_404(
                    MCQBank,
                    id=q_id,
                    is_active=True
                )

                ExamMCQ.objects.create(
                    exam=exam,
                    source_type="bank",
                    bank_question=bank_q,
                    question_text=bank_q.question_text,
                    option_a=bank_q.option_a,
                    option_b=bank_q.option_b,
                    option_c=bank_q.option_c,
                    option_d=bank_q.option_d,
                    correct_option=bank_q.correct_option,
                    marks=1,
                    order=exam.mcqs.count() + 1
                )

            return redirect(request.path)

        # --------------------------------
        # ADD SHORT ANSWER FROM BANK
        # --------------------------------
        elif action == "add_bank_short":

            selected_ids = request.POST.getlist("selected_shorts")

            for q_id in selected_ids:
                bank_q = get_object_or_404(
                    ShortAnswerBank,
                    id=q_id,
                    is_active=True
                )

                ExamShortAnswer.objects.create(
                    exam=exam,
                    source_type="bank",
                    bank_question=bank_q,
                    question_text=bank_q.question_text,
                    marks=5,
                    order=exam.short_answers.count() + 1
                )

            return redirect(request.path)

        # --------------------------------
        # UPDATE MCQ MARKS
        # --------------------------------
        elif action == "update_mcq_marks":

            mcq_id = request.POST.get("mcq_id")
            marks = request.POST.get("marks")

            mcq = get_object_or_404(
                ExamMCQ,
                id=mcq_id,
                exam=exam
            )

            mcq.marks = int(marks)
            mcq.save()

            return redirect(request.path)

        # --------------------------------
        # UPDATE SHORT MARKS
        # --------------------------------
        elif action == "update_short_marks":

            short_id = request.POST.get("short_id")
            marks = request.POST.get("marks")

            short = get_object_or_404(
                ExamShortAnswer,
                id=short_id,
                exam=exam
            )

            short.marks = int(marks)
            short.save()

            return redirect(request.path)
        # 🔥 ADD THIS BLOCK RIGHT HERE
        elif action == "update_viva_marks":

            viva_marks = request.POST.get("viva_marks")

            exam.viva_marks = int(viva_marks)
            exam.save()

            return redirect(request.path)

        # --------------------------------
        # UPDATE PRACTICAL RECORD MARKS
        # --------------------------------
        elif action == "update_practical_record_marks":
            
            pr_marks = request.POST.get("practical_record_marks")
            exam.practical_record_marks = int(pr_marks)
            exam.save()
            
            return redirect(request.path)
# --------------------------------
# ADD MAJOR / MINOR PRACTICAL
# --------------------------------
        elif action == "add_practical":

            practical_type = request.POST.get("practical_type")
            experiment_id = request.POST.get("experiment_id")
            marks = request.POST.get("marks")

            selected_experiment = get_object_or_404(
                Experiment,
                id=experiment_id,
                is_active=True
            )
        # elif action == "update_viva_marks":

        #     viva_marks = request.POST.get("viva_marks")

        #     exam.viva_marks = int(viva_marks)
        #     exam.save()

        #     return redirect(request.path)

            # Ensure only one major and one minor per exam
            if ExamPractical.objects.filter(
                exam=exam,
                practical_type=practical_type
            ).exists():
                return redirect(request.path)

            ExamPractical.objects.create(
                exam=exam,
                practical_type=practical_type,
                experiment=selected_experiment,
                title=selected_experiment.name,
                aim=selected_experiment.aim,
                marks=int(marks)
            )

            return redirect(request.path)
    # ================================
    # FINAL RENDER
    # ================================
    major_practical = exam.practicals.filter(practical_type="major").first()
    minor_practical = exam.practicals.filter(practical_type="minor").first()

    # 🔥 Filter experiments based on field (D.Pharm vs B.Pharm)
    experiments_list = Experiment.objects.filter(is_active=True).order_by("number")
    
    # We pass the full list to template, and let the template handle 
    # D.Pharm re-indexing/filtering just like in assign_practical
    is_dpharm_exam = (exam.year == 'dpharm_2')

    # 📂 Organize into sections for the searchable dropdown
    builder_sections = [
        {"id": "A", "title": "Introduction & Basic Concepts", "exps": []},
        {"id": "B", "title": "Ethics, Handling & Techniques", "exps": []},
        {"id": "C", "title": "Autonomic & Peripheral Pharmacology", "exps": []},
        {"id": "D", "title": "Central Nervous System (CNS)", "exps": []},
        {"id": "E", "title": "Cardiovascular & Renal Pharmacology", "exps": []},
        {"id": "F", "title": "Analgesic & Anti-Inflammatory", "exps": []},
        {"id": "G", "title": "Endocrine & Metabolic Pharmacology", "exps": []},
        {"id": "H", "title": "Gastrointestinal & Anti-Ulcer", "exps": []},
        {"id": "I", "title": "Anaesthetics & Allergy Models", "exps": []},
        {"id": "J", "title": "In-Vitro Bioassay Studies", "exps": []},
        {"id": "K", "title": "Toxicology Studies", "exps": []},
        {"id": "L", "title": "Biomechanics & Biostatistics", "exps": []},
    ]

    for exp in experiments_list:
        num = exp.number
        if 1 <= num <= 6: builder_sections[0]["exps"].append(exp)
        elif 7 <= num <= 9: builder_sections[1]["exps"].append(exp)
        elif 10 <= num <= 15: builder_sections[2]["exps"].append(exp)
        elif 16 <= num <= 22: builder_sections[3]["exps"].append(exp)
        elif 23 <= num <= 29: builder_sections[4]["exps"].append(exp)
        elif 30 <= num <= 33: builder_sections[5]["exps"].append(exp)
        elif num == 34: builder_sections[6]["exps"].append(exp)
        elif num == 35: builder_sections[7]["exps"].append(exp)
        elif 36 <= num <= 40: builder_sections[8]["exps"].append(exp)
        elif 41 <= num <= 50: builder_sections[9]["exps"].append(exp)
        elif 51 <= num <= 53: builder_sections[10]["exps"].append(exp)
        elif 54 <= num <= 56: builder_sections[11]["exps"].append(exp)

    # Get MCQ and Short Answer banks linked to filtered experiments
    # (Or all active questions if not linked)
    mcq_bank = MCQBank.objects.filter(is_active=True).order_by('experiment__name', 'question_text')
    short_bank = ShortAnswerBank.objects.filter(is_active=True).order_by('experiment__name', 'question_text')

    return render(
        request,
        "teacher/exam_builder.html",
        {
            "exam": exam,
            "mcq_bank": mcq_bank,
            "short_bank": short_bank,
            "mcqs": exam.mcqs.all(),
            "short_answers": exam.short_answers.all(),
            "spotting_questions": exam.spotting_questions.all(),
            "spotting_bank": SpottingBank.objects.filter(is_active=True).order_by("name"),
            "practicals": exam.practicals.all(),
            "experiments_list": experiments_list,
            "major_practical": major_practical,
            "minor_practical": minor_practical,
            "builder_sections": builder_sections,
            "is_dpharm_exam": is_dpharm_exam,
        }
    )
    

def delete_exam_mcq(request, mcq_id):
    if request.session.get("role") != "teacher":
        return redirect("/login/")

    mcq = get_object_or_404(
        ExamMCQ,
        id=mcq_id,
        exam__teacher=request.user
    )

    exam_id = mcq.exam.id
    mcq.delete()

    return redirect(f"/teacher/exams/{exam_id}/builder/")
def delete_exam_short(request, short_id):
    if request.session.get("role") != "teacher":
        return redirect("/login/")

    short = get_object_or_404(
        ExamShortAnswer,
        id=short_id,
        exam__teacher=request.user
    )

    exam_id = short.exam.id
    short.delete()

    return redirect(f"/teacher/exams/{exam_id}/builder/")

def delete_exam_spotting(request, spot_id):
    if request.session.get("role") != "teacher":
        return redirect("/login/")

    spot = get_object_or_404(
        ExamSpotting,
        id=spot_id,
        exam__teacher=request.user
    )

    exam_id = spot.exam.id
    spot.delete()

    return redirect(f"/teacher/exams/{exam_id}/builder/")


# =====================================================
# 🔥 TEACHER → VIEW EXAM ATTEMPTS
# =====================================================
from .models import ExamAttempt

def teacher_exam_attempts(request, exam_id):

    if request.session.get("role") != "teacher":
        return redirect("/login/")

    exam = get_object_or_404(
        Exam,
        id=exam_id,
        teacher=request.user
    )

    attempts = ExamAttempt.objects.filter(
        exam=exam,
        status__in=["submitted", "approved"]
    ).select_related("student")

    return render(
        request,
        "teacher/exam_attempts.html",
        {
            "exam": exam,
            "attempts": attempts
        }
    )

def export_exam_submissions_pdf(request, exam_id):
    if request.session.get("role") != "teacher":
        return redirect("/login/")

    exam = get_object_or_404(Exam, id=exam_id, teacher=request.user)
    college = request.user.college
    teacher = request.user

    attempts = ExamAttempt.objects.filter(
        exam=exam,
        status__in=["submitted", "approved"]
    ).select_related("student").order_by("student__roll_no")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=50)
    elements = []
    styles = getSampleStyleSheet()

    # --- Header (Logo and Name) ---
    college_name = college.name if college else "Virtual Lab System"
    college_address = college.address if college and college.address else ""
    
    logo_part = None
    if college and college.logo:
        try:
            logo_path = college.logo.path
            if os.path.exists(logo_path):
                logo_part = Image(logo_path, width=65, height=65)
        except Exception:
            pass

    name_style = ParagraphStyle(
        'CollegeNameStyle',
        parent=styles['Normal'],
        fontName='Times-Bold',
        fontSize=22,
        textColor=colors.black,
        alignment=0, # Left
        leading=24
    )

    address_style = ParagraphStyle(
        'AddressStyle',
        parent=styles['Normal'],
        fontName='Times-Roman',
        fontSize=9,
        textColor=colors.grey,
        alignment=0, # Left
    )
    
    header_info = [Paragraph(college_name, name_style)]
    if college_address:
        header_info.append(Paragraph(college_address, address_style))
    
    # Teacher Info
    header_info.append(Paragraph(f"Teacher: {teacher.first_name} {teacher.last_name}", address_style))

    if logo_part:
        header_table = Table([[logo_part, header_info]], colWidths=[80, 450])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ]))
        elements.append(header_table)
    else:
        elements.extend(header_info)
    
    elements.append(Spacer(1, 15))
    elements.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor("#ff4b2b"), spaceAfter=15))
    
    # Title
    title_style = ParagraphStyle(
        'MainTitleStyle',
        parent=styles['Normal'],
        fontName='Times-Bold',
        fontSize=16,
        alignment=1, # Center
        spaceAfter=15,
    )
    elements.append(Paragraph(f"<u>SUBMISSION LIST: {exam.title.upper()}</u>", title_style))
    
    type_style = ParagraphStyle(
        'TypeStyle',
        parent=styles['Normal'],
        fontName='Times-Bold',
        fontSize=11,
        alignment=1,
        spaceAfter=10,
        textColor=colors.HexColor("#6366f1")
    )
    elements.append(Paragraph(f"({exam.get_exam_type_display()})", type_style))

    # --- Table Data ---
    if exam.exam_type == "external":
        data = [["Roll No", "Seat No", "Student Name", "Status", "Final Marks"]]
        col_widths = [65, 80, 180, 90, 110]
    else:
        data = [["Roll No", "Student Name", "Status", "Final Marks"]]
        col_widths = [80, 220, 100, 125]

    for att in attempts:
        status = "Evaluated" if att.status == "approved" else "Pending"
        score = f"{att.total_score} / {exam.total_max_marks}" if att.status == "approved" else "---"
        
        row = [att.student.roll_no or "---"]
        if exam.exam_type == "external":
            row.append(att.seat_number or "---")
        
        row.extend([
            f"{att.student.first_name} {att.student.last_name}",
            status,
            score
        ])
        data.append(row)

    table = Table(data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#ff4b2b")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Times-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, 1), (-1, -1), 'Times-Roman'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#fff5f5")]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(table)

    def page_setup(canvas, doc):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor("#ff4b2b"))
        canvas.setLineWidth(1)
        canvas.rect(20, 20, A4[0]-40, A4[1]-40)
        
        try:
            gmars_logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'logo.png')
            if os.path.exists(gmars_logo_path):
                canvas.drawImage(gmars_logo_path, A4[0] - 110, 10, width=80, height=80, mask='auto')
        except Exception:
            pass
        
        canvas.setFont('Times-BoldItalic', 10)
        canvas.drawRightString(A4[0] - 120, 40, "Signature of Examiner")
        canvas.setFont('Times-Roman', 8)
        canvas.drawString(40, 40, f"Generated on: {now().strftime('%d-%m-%Y %H:%M')}")
        canvas.drawCentredString(A4[0]/2, 40, f"Page {canvas.getPageNumber()}")
        canvas.restoreState()

    try:
        doc.build(elements, onFirstPage=page_setup, onLaterPages=page_setup)
    except Exception as e:
        return HttpResponse(f"Error: {str(e)}", status=500)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Submissions_{exam.id}.pdf"'
    response.write(buffer.getvalue())
    buffer.close()
    return response


# =====================================================
# 🔥 TEACHER → EVALUATE ATTEMPT
# =====================================================
def render_student_table(html_string, student_table_data):
    """
    Replaces empty <td></td> with <td>STUDENT_VALUE</td>
    """
    if not html_string:
        return ""
    
    def replacer(match):
        nonlocal cell_count
        cell_count += 1
        val = student_table_data.get(str(cell_count), "")
        return f'<td style="background:rgba(99, 102, 241, 0.08); color:#4f46e5; font-weight:700; border: 1px solid rgba(99,102,241,0.2); text-align:center; padding:12px;">{val}</td>'
    
    cell_count = 0
    modified_html = re.sub(r'<td>\s*</td>', replacer, html_string)
    return modified_html

@csrf_exempt
def evaluate_attempt(request, attempt_id):

    if request.session.get("role") != "teacher":
        return redirect("/login/")

    attempt = get_object_or_404(
        ExamAttempt,
        id=attempt_id,
        exam__teacher=request.user
    )
    
    exam = attempt.exam
    is_dpharm = exam.year == "dpharm_2"
    answers = attempt.answers or {}

    if request.method == "POST":

        # 🔥 MANUAL SCORES FOR ALL SECTIONS
        spotting_score = float(request.POST.get("spotting_score", 0))
        mcq_score = float(request.POST.get("mcq_score", 0))
        short_score = float(request.POST.get("short_score", 0))
        practical_score = float(request.POST.get("practical_score", 0))
        viva_score = float(request.POST.get("viva_score", 0))
        practical_record_score = float(request.POST.get("practical_record_score", 0))

        attempt.spotting_score = spotting_score
        attempt.mcq_score = mcq_score
        attempt.short_score = short_score
        attempt.practical_score = practical_score
        attempt.viva_score = viva_score
        attempt.practical_record_score = practical_record_score

        # 🔥 Final total calculation
        attempt.total_score = (
            spotting_score +
            mcq_score +
            short_score +
            practical_score +
            viva_score +
            practical_record_score
        )

        attempt.status = "approved"
        attempt.teacher_approved = True
        attempt.save()

        # Redirect to exam attempts list
        return redirect(f"/teacher/exams/{exam.id}/attempts/")

    # 🔥 PRE-PROCESS PRACTICAL TABLES FOR DISPLAY
    practicals_data = []
    for p in exam.practicals.all():
        student_prac = answers.get("practical", {}).get(str(p.id), {})
        table_html = EXPERIMENT_TABLES.get(p.experiment.slug)
        rendered_table = render_student_table(table_html, student_prac.get("table", {}))
        
        practicals_data.append({
            "obj": p,
            "student_data": student_prac,
            "rendered_table": rendered_table
        })

    return render(
        request,
        "teacher/evaluate_attempt.html",
        {
            "attempt": attempt,
            "answers": answers,
            "is_dpharm": is_dpharm,
            "exam": exam,
            "practicals_data": practicals_data,
        }
    )

def delete_exam_practical(request, practical_id):
    if request.session.get("role") != "teacher":
        return redirect("/login/")

    practical = get_object_or_404(
        ExamPractical,
        id=practical_id,
        exam__teacher=request.user
    )

    exam_id = practical.exam.id
    practical.delete()

    return redirect(f"/teacher/exams/{exam_id}/builder/")

# =====================================================
# 🔥 SESSIONAL MARKSHEET: SELECTION PAGE
# =====================================================
def sessional_marksheet_select(request):
    if request.session.get("role") != "teacher":
        return redirect("/login/")

    teacher = request.user
    subject_code = request.GET.get("subject")
    
    # Selection 1: Internal 1
    exams_1 = Exam.objects.filter(
        teacher=teacher,
        exam_type="internal_1"
    )
    if subject_code:
        exams_1 = exams_1.filter(year=subject_code)
    exams_1 = exams_1.order_by("-created_at")

    # Selection 2: Internal 2
    exams_2 = Exam.objects.filter(
        teacher=teacher,
        exam_type="internal_2"
    )
    if subject_code:
        exams_2 = exams_2.filter(year=subject_code)
    exams_2 = exams_2.order_by("-created_at")

    return render(
        request,
        "teacher/sessional_select.html",
        {
            "exams_1": exams_1,
            "exams_2": exams_2,
            "subject_code": subject_code
        }
    )

# =====================================================
# 🔥 SESSIONAL MARKSHEET: MAIN GRID VIEW
# =====================================================
def sessional_marksheet_view(request):
    if request.session.get("role") != "teacher":
        return redirect("/login/")

    exam_1_id = request.GET.get("exam_1")
    exam_2_id = request.GET.get("exam_2")

    if not exam_1_id or not exam_2_id:
        return redirect("sessional-marksheet-select")

    teacher = request.user
    ex1 = get_object_or_404(Exam, id=exam_1_id, teacher=teacher)
    ex2 = get_object_or_404(Exam, id=exam_2_id, teacher=teacher)

    # Students assigned to this teacher/college AND matched to this subject/year
    students = User.objects.filter(
        role="student",
        subject=ex1.year,
        created_by=teacher,
        college=teacher.college
    ).order_by("roll_no")

    # Fetch all attempts for these exams at once to avoid N+1
    attempts_1 = {a.student_id: a for a in ExamAttempt.objects.filter(exam=ex1, status="approved")}
    attempts_2 = {a.student_id: a for a in ExamAttempt.objects.filter(exam=ex2, status="approved")}

    # Fetch continuous marks
    continuous_marks = {m.student_id: m.continuous_score for m in SessionalContinuousMark.objects.filter(
        teacher=teacher, 
        year=ex1.year,
        subject_code=ex1.title
    )}

    marksheet_data = []

    for s in students:
        mark_1 = attempts_1.get(s.id).total_score if s.id in attempts_1 else 0
        mark_2 = attempts_2.get(s.id).total_score if s.id in attempts_2 else 0
        cont_mark = continuous_marks.get(s.id, 0)

        # Average
        avg = (mark_1 + mark_2) / 2
        
        # Converted (Out of 10)
        total_max = (ex1.total_max_marks + ex2.total_max_marks) / 2
        converted = (avg / total_max * 10) if total_max > 0 else 0
        
        final_total = cont_mark + converted

        marksheet_data.append({
            "student": s,
            "mark_1": mark_1,
            "mark_2": mark_2,
            "continuous": cont_mark,
            "average": round(avg, 2),
            "converted": round(converted, 2),
            "final_total": round(final_total, 2)
        })

    return render(
        request,
        "teacher/sessional_view.html",
        {
            "ex1": ex1,
            "ex2": ex2,
            "data": marksheet_data
        }
    )

# =====================================================
# 🔥 AJAX: SAVE CONTINUOUS MARK
# =====================================================
@csrf_exempt
def save_continuous_mark(request):
    if request.method != "POST" or request.session.get("role") != "teacher":
        return JsonResponse({"status": "error", "message": "Unauthorized"}, status=403)

    student_id = request.POST.get("student_id")
    year = request.POST.get("year")
    subject_code = request.POST.get("subject_code")
    score = request.POST.get("score")

    student = get_object_or_404(User, id=student_id, role="student")
    
    obj, created = SessionalContinuousMark.objects.update_or_create(
        student=student,
        year=year,
        subject_code=subject_code,
        defaults={
            "teacher": request.user,
            "continuous_score": float(score or 0)
        }
    )

    return JsonResponse({
        "status": "success", 
        "continuous_score": obj.continuous_score,
        "message": "Saved successfully"
    })

# =====================================================
# 🔥 EXPORT SESSIONAL MARKSHEET → PDF
# =====================================================
def export_sessional_marksheet_pdf(request):
    if request.session.get("role") != "teacher":
        return redirect("/login/")

    exam_1_id = request.GET.get("ex1")
    exam_2_id = request.GET.get("ex2")

    ex1 = get_object_or_404(Exam, id=exam_1_id, teacher=request.user)
    ex2 = get_object_or_404(Exam, id=exam_2_id, teacher=request.user)
    college = request.user.college
    teacher = request.user

    # Filter students strictly by the exam's year/subject
    students = User.objects.filter(
        role="student", 
        subject=ex1.year,
        created_by=teacher, 
        college=college
    ).order_by("roll_no")

    attempts_1 = {a.student_id: a for a in ExamAttempt.objects.filter(exam=ex1, status="approved")}
    attempts_2 = {a.student_id: a for a in ExamAttempt.objects.filter(exam=ex2, status="approved")}
    continuous_marks = {m.student_id: m.continuous_score for m in SessionalContinuousMark.objects.filter(
        teacher=teacher, year=ex1.year, subject_code=ex1.title
    )}

    buffer = io.BytesIO()
    # Margins matched to export_exam_submissions_pdf
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=50)
    elements = []
    styles = getSampleStyleSheet()

    # --- Header (Logo and Name) ---
    college_name = college.name if college else "Virtual Lab System"
    college_address = college.address if college and college.address else ""
    logo_part = None
    if college and college.logo:
        try:
            logo_path = college.logo.path
            if os.path.exists(logo_path):
                logo_part = Image(logo_path, width=65, height=65)
        except Exception: pass

    name_style = ParagraphStyle('CN', parent=styles['Normal'], fontName='Times-Bold', fontSize=22, alignment=0, leading=24)
    addr_style = ParagraphStyle('AD', parent=styles['Normal'], fontName='Times-Roman', fontSize=9, textColor=colors.grey, alignment=0)
    
    h_info = [Paragraph(college_name, name_style)]
    if college_address: h_info.append(Paragraph(college_address, addr_style))
    h_info.append(Paragraph(f"Teacher: {teacher.first_name} {teacher.last_name}", addr_style))

    if logo_part:
        header_table = Table([[logo_part, h_info]], colWidths=[80, 450])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ]))
        elements.append(header_table)
    else:
        elements.extend(h_info)

    elements.append(Spacer(1, 15))
    elements.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor("#ff4b2b"), spaceAfter=15))

    # Title
    t_style = ParagraphStyle('TS', parent=styles['Normal'], fontName='Times-Bold', fontSize=14, alignment=1, spaceAfter=20)
    elements.append(Paragraph(f"<u>SESSIONAL MARKSHEET - {ex1.get_year_display().upper()}</u>", t_style))
    
    sub_style = ParagraphStyle('SUB', parent=styles['Normal'], fontName='Times-Bold', fontSize=11, alignment=1, spaceAfter=15, textColor=colors.HexColor("#6366f1"))
    elements.append(Paragraph(f"Subject: {ex1.title}", sub_style))

    # --- Table Data ---
    headers = ["Roll No", "Student Name", "Continuous Mode\n(5 Marks)", "Sess. 1", "Sess. 2", "Avg", "Conversion(10)", "Total"]
    data = [headers]

    for s in students:
        m1 = attempts_1.get(s.id).total_score if s.id in attempts_1 else 0
        m2 = attempts_2.get(s.id).total_score if s.id in attempts_2 else 0
        cont = continuous_marks.get(s.id, 0)
        avg = (m1 + m2) / 2
        t_max = (ex1.total_max_marks + ex2.total_max_marks) / 2
        conv = (avg / t_max * 10) if t_max > 0 else 0
        total = cont + conv

        data.append([
            s.roll_no or "-",
            f"{s.first_name} {s.last_name}",
            f"{cont}",
            f"{m1}",
            f"{m2}",
            f"{round(avg, 1)}",
            f"{round(conv, 1)}",
            f"{round(total, 1)}"
        ])

    table = Table(data, colWidths=[35, 145, 90, 40, 40, 40, 65, 40])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#ff4b2b")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Times-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 0), (-1, 0), 8.5),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#fff5f5")]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
    ]))
    elements.append(table)

    def page_setup(canvas, doc):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor("#ff4b2b"))
        canvas.setLineWidth(1)
        # Rect matched to submission PDF
        canvas.rect(20, 20, A4[0]-40, A4[1]-40)
        
        try:
            gmars_logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'logo.png')
            if os.path.exists(gmars_logo_path):
                # Placement matched to submission PDF
                canvas.drawImage(gmars_logo_path, A4[0] - 110, 10, width=80, height=80, mask='auto')
        except: pass
        
        canvas.setFont('Times-BoldItalic', 10)
        canvas.drawRightString(A4[0] - 120, 40, "Signature of Examiner")
        canvas.setFont('Times-Roman', 8)
        canvas.drawString(40, 40, f"Generated on: {now().strftime('%d-%m-%Y %H:%M')}")
        canvas.drawCentredString(A4[0]/2, 40, f"Page {canvas.getPageNumber()}")
        canvas.restoreState()

    try:
        doc.build(elements, onFirstPage=page_setup, onLaterPages=page_setup)
    except Exception as e: return HttpResponse(f"Error: {str(e)}", status=500)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Sessional_Marksheet.pdf"'
    response.write(buffer.getvalue())
    buffer.close()
    return response